import os, warnings, json
warnings.filterwarnings('ignore')
import openpyxl

folder = r'c:\Users\X1 Yoga\OneDrive\Desktop\New folder'
wb = openpyxl.load_workbook(os.path.join(folder, 'Data Vận hành.xlsx'), data_only=True)

# Extract full data from Shopee sheet (which has volume data by district/buu cuc)
ws = wb['Shopee']
print('=== Shopee sheet full headers ===')
for i in range(1,5):
    for j in range(1, ws.max_column+1):
        v = ws.cell(i,j).value
        if v: print(f'  R{i}C{j}: {v}')

print('\n=== Shopee sample 5-12 ===')
for i in range(5, min(13, ws.max_row+1)):
    for j in range(1, ws.max_column+1):
        v = ws.cell(i,j).value
        if v: print(f'  R{i}C{j}: {v}', end=' | ')
    print()

# Count non-empty data rows
shopee_count = sum(1 for i in range(5, ws.max_row+1) if ws.cell(i,18).value)
print(f'Shopee data rows: {shopee_count}')

# Get all unique values from Shopee 
shopee_ams = set()
shopee_provinces = set()
for i in range(5, ws.max_row+1):
    am = ws.cell(i,20).value
    prov = ws.cell(i,18).value
    if am: shopee_ams.add(am)
    if prov: shopee_provinces.add(prov)
print(f'Shopee AMs: {sorted(shopee_ams)}')
print(f'Shopee provinces: {sorted(shopee_provinces)}')

# Also extract Tất cả sheet full data for export
ws2 = wb['Tất cả']
all_data = []
for i in range(5, ws2.max_row+1):
    row = {}
    stt = ws2.cell(i,1).value
    prov = ws2.cell(i,18).value
    dist = ws2.cell(i,19).value
    am = ws2.cell(i,20).value
    cap = ws2.cell(i,21).value
    vol = ws2.cell(i,22).value
    loai = ws2.cell(i,23).value
    if prov:
        all_data.append({
            'stt': stt, 'province': prov, 'district': dist,
            'am': am, 'capacity': cap, 'volume': vol, 'type': loai
        })
print(f'\nTất cả data: {len(all_data)} rows')
# Save as JSON for web app
with open(os.path.join(folder, 'data_export.json'), 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)
print('Exported to data_export.json')

# Extract Cơ cấu data
ws3 = wb['Cơ cấu']
cocau_data = []
for i in range(2, ws3.max_row+1):
    wh_id = ws3.cell(i,1).value
    bc = ws3.cell(i,2).value
    ward = ws3.cell(i,4).value
    dist = ws3.cell(i,6).value
    prov = ws3.cell(i,8).value
    am = ws3.cell(i,9).value
    if bc:
        cocau_data.append({
            'warehouse_id': wh_id, 'buu_cuc': bc, 'ward': ward,
            'district': dist, 'province': prov, 'am': am
        })
print(f'Cơ cấu data: {len(cocau_data)} rows')
with open(os.path.join(folder, 'cocau_export.json'), 'w', encoding='utf-8') as f:
    json.dump(cocau_data, f, ensure_ascii=False, indent=2)
print('Exported to cocau_export.json')

# Also extract Shopee data
ws4 = wb['Shopee']
shopee_data = []
for i in range(5, ws4.max_row+1):
    prov = ws4.cell(i,18).value
    dist = ws4.cell(i,19).value
    am = ws4.cell(i,20).value
    if prov:
        row = {'province': prov, 'district': dist, 'am': am}
        # Get volume columns (C4-C17 seem to be daily volumes)
        for j in range(2, 18):
            v = ws4.cell(i,j).value
            if v:
                header = ws4.cell(4,j).value or f'col{j}'
                row[str(header)] = v
        shopee_data.append(row)
print(f'Shopee data: {len(shopee_data)} rows')
with open(os.path.join(folder, 'shopee_export.json'), 'w', encoding='utf-8') as f:
    json.dump(shopee_data, f, ensure_ascii=False, indent=2)
print('Exported to shopee_export.json')
