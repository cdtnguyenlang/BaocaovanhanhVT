import os, warnings, json
from datetime import datetime, date, time as dt_time
warnings.filterwarnings('ignore')
import openpyxl

folder = r'c:\Users\X1 Yoga\OneDrive\Desktop\New folder'
wb = openpyxl.load_workbook(os.path.join(folder, 'Data Vận hành.xlsx'), data_only=True)

def ser(val):
    """Serialize Excel cell value to JSON-safe format"""
    if val is None: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, date): return val.strftime('%Y-%m-%d')
    if isinstance(val, dt_time): return val.strftime('%H:%M')
    if isinstance(val, (int, float)): return val
    s = str(val).strip()
    if s in ('#N/A','#DIV/0!','#VALUE!','#REF!','#NAME?',''): return None
    return s

# ============ 1. THÔNG TIN XE ============
ws = wb['Thông tin xe']
vehicles = []
for i in range(2, ws.max_row+1):
    plate = ws.cell(i,2).value
    if not plate: continue
    vehicles.append({
        'stt': ser(ws.cell(i,1).value),
        'plate': ser(plate),
        'tonnage': ser(ws.cell(i,3).value),
        'model': ser(ws.cell(i,4).value),
        'region': ser(ws.cell(i,6).value),
        'department': ser(ws.cell(i,7).value),
        'boxVolume': ser(ws.cell(i,8).value),
        'yearReceived': ser(ws.cell(i,10).value),
        'yearsUsed': ser(ws.cell(i,12).value),
        'condition': ser(ws.cell(i,13).value),
        'status': ser(ws.cell(i,14).value),
        'insuranceExpiry': ser(ws.cell(i,15).value),
        'inspectionCode': ser(ws.cell(i,16).value),
        'inspectionExpiry': ser(ws.cell(i,17).value),
        'liabilityExpiry': ser(ws.cell(i,18).value),
        'roadFeeExpiry': ser(ws.cell(i,19).value),
        'badgeExpiry': ser(ws.cell(i,20).value),
        'regCertExpiry': ser(ws.cell(i,21).value),
        'totalKm': ser(ws.cell(i,24).value),
        'warning': ser(ws.cell(i,26).value),
        'note': ser(ws.cell(i,27).value),
        'fleet': ser(ws.cell(i,28).value),
    })

# ============ 2. LỊCH TẢI ============
ws = wb['Lịch tải']
routes = []
for i in range(2, ws.max_row+1):
    rname = ws.cell(i,1).value
    if not rname: continue
    routes.append({
        'routeName': ser(rname),
        'tonnage': ser(ws.cell(i,2).value),
        'warehouse': ser(ws.cell(i,3).value),
        'type': ser(ws.cell(i,4).value),
        'arrival': ser(ws.cell(i,5).value),
        'departure': ser(ws.cell(i,6).value),
        'note': ser(ws.cell(i,7).value),
        'km': ser(ws.cell(i,8).value),
        'supplier': ser(ws.cell(i,9).value),
    })

# ============ 3. PHẠT NGUỘI ============
ws = wb['Phạt nguội']
fines = []
for i in range(2, ws.max_row+1):
    plate = ws.cell(i,4).value
    if not plate: continue
    fines.append({
        'reportDate': ser(ws.cell(i,1).value),
        'plate': ser(plate),
        'depot': ser(ws.cell(i,5).value),
        'violationTime': ser(ws.cell(i,6).value),
        'location': ser(ws.cell(i,7).value),
        'violation': ser(ws.cell(i,8).value),
        'cost': ser(ws.cell(i,9).value),
        'sup': ser(ws.cell(i,11).value),
        'driverId': ser(ws.cell(i,12).value),
        'driverName': ser(ws.cell(i,13).value),
        'driverStatus': ser(ws.cell(i,14).value),
        'expectedDate': ser(ws.cell(i,15).value),
        'progress': ser(ws.cell(i,17).value),
    })

# ============ 4. HIỆU SUẤT SỬ DỤNG XE ============
ws = wb['Hiệu suất sử dụng xe']
efficiency = []
for i in range(2, ws.max_row+1):
    plate = ws.cell(i,2).value
    if not plate: continue
    eff_val = ws.cell(i,16).value
    efficiency.append({
        'stt': ser(ws.cell(i,1).value),
        'plate': ser(plate),
        'tonnage': ser(ws.cell(i,3).value),
        'model': ser(ws.cell(i,4).value),
        'region': ser(ws.cell(i,6).value),
        'department': ser(ws.cell(i,7).value),
        'yearsUsed': ser(ws.cell(i,12).value),
        'condition': ser(ws.cell(i,13).value),
        'status': ser(ws.cell(i,14).value),
        'vehicleType': ser(ws.cell(i,15).value),
        'efficiency': round(float(eff_val)*100, 1) if isinstance(eff_val, (int,float)) else 0,
        'opStatus': ser(ws.cell(i,17).value),
    })

# ============ 5. TÀI XẾ ============
ws = wb['Tài xế']
drivers = []
for i in range(2, ws.max_row+1):
    name = ws.cell(i,3).value
    if not name: continue
    drivers.append({
        'stt': ser(ws.cell(i,1).value),
        'employeeId': ser(ws.cell(i,2).value),
        'name': ser(name),
        'phone': ser(ws.cell(i,4).value),
        'position': ser(ws.cell(i,5).value),
        'unit': ser(ws.cell(i,6).value),
        'supervisor': ser(ws.cell(i,7).value),
        'shift': ser(ws.cell(i,8).value),
        'route': ser(ws.cell(i,9).value),
        'startDate': ser(ws.cell(i,10).value),
        'endDate': ser(ws.cell(i,11).value),
        'status': ser(ws.cell(i,12).value),
        'seniority': ser(ws.cell(i,13).value),
        'seniorityDetail': ser(ws.cell(i,14).value),
    })

# ============ 6. TẢI TĂNG CƯỜNG LẤY ============
ws = wb['Tải tăng cường Lấy']
reinforcement = []
for i in range(2, ws.max_row+1):
    tid = ws.cell(i,1).value
    if not tid: continue
    reinforcement.append({
        'ticketId': ser(tid),
        'region': ser(ws.cell(i,2).value),
        'warehouse': ser(ws.cell(i,3).value),
        'route': ser(ws.cell(i,5).value),
        'employeeId': ser(ws.cell(i,6).value),
        'phone': ser(ws.cell(i,8).value),
        'packages': ser(ws.cell(i,9).value),
        'volumeNeeded': ser(ws.cell(i,10).value),
        'requestDate': ser(ws.cell(i,11).value),
        'note': ser(ws.cell(i,12).value),
        'status': ser(ws.cell(i,13).value),
        'date': ser(ws.cell(i,14).value),
        'arrivalTime': ser(ws.cell(i,15).value),
        'tripCode': ser(ws.cell(i,16).value),
        'supplier': ser(ws.cell(i,17).value),
        'plate': ser(ws.cell(i,18).value),
        'tonnage': ser(ws.cell(i,19).value),
        'driverInfo': ser(ws.cell(i,20).value),
    })

# ============ BUILD data.js ============
data_js = f"""// GHN Operations Dashboard - Data Module
// Auto-generated from Data Vận hành.xlsx on {datetime.now().strftime('%Y-%m-%d %H:%M')}

const DATA = {{
  lastUpdated: '{datetime.now().strftime('%Y-%m-%d')}',
  region: 'Vùng Hồ Chí Minh',

  vehicles: {json.dumps(vehicles, ensure_ascii=False, indent=2)},

  routes: {json.dumps(routes, ensure_ascii=False, indent=2)},

  fines: {json.dumps(fines, ensure_ascii=False, indent=2)},

  efficiency: {json.dumps(efficiency, ensure_ascii=False, indent=2)},

  drivers: {json.dumps(drivers, ensure_ascii=False, indent=2)},

  reinforcement: {json.dumps(reinforcement, ensure_ascii=False, indent=2)},

  companyInfo: {{
    name: 'Giao Hàng Nhanh (GHN)',
    parent: 'SCOMMERCE',
    fleet: '1,100+ xe tải',
    postOffices: '1,000+ bưu cục',
    sortingCenters: '30+ trung tâm phân loại',
    deliveryStaff: '20,000+ nhân viên',
    dailyCapacity: '5.5 triệu đơn/ngày',
    hcmDailyVolume: '300,000+ đơn/ngày'
  }}
}};
"""

with open(os.path.join(folder, 'data.js'), 'w', encoding='utf-8') as f:
    f.write(data_js)

print(f'data.js generated successfully!')
print(f'  Vehicles: {len(vehicles)}')
print(f'  Routes: {len(routes)}')
print(f'  Fines: {len(fines)}')
print(f'  Efficiency: {len(efficiency)}')
print(f'  Drivers: {len(drivers)}')
print(f'  Reinforcement: {len(reinforcement)}')
