import os, warnings
warnings.filterwarnings('ignore')
import openpyxl

folder = r'c:\Users\X1 Yoga\OneDrive\Desktop\New folder'
wb = openpyxl.load_workbook(os.path.join(folder, 'Data Vận hành.xlsx'), data_only=True)

# Check the Shopee sheet with volume data (has C2=district, C3=buu_cuc, C4-C17=volumes)
for sname in ['Shopee', 'Tất cả']:
    ws = wb[sname]
    print(f'\n=== {sname}: rows={ws.max_row} ===')
    # Get headers from rows 1-4
    for i in range(1,5):
        row_data = []
        for j in range(1, ws.max_column+1):
            v = ws.cell(i,j).value
            if v: row_data.append(f'C{j}={v}')
        if row_data: print(f'  H-R{i}: {row_data}')
    
    # Sample data
    for i in range(5, min(12, ws.max_row+1)):
        row_data = []
        for j in range(1, ws.max_column+1):
            v = ws.cell(i,j).value
            if v: row_data.append(f'C{j}={v}')
        if row_data: print(f'  D-R{i}: {row_data}')

    # Summary stats
    all_data = []
    for i in range(5, ws.max_row+1):
        row = {}
        for j in range(1, ws.max_column+1):
            v = ws.cell(i,j).value
            if v: row[f'C{j}'] = v
        if row: all_data.append(row)
    print(f'  Total data rows with content: {len(all_data)}')

# Also check TP.HCM and Theo tinh sheets more deeply
for sname in ['TP. HCM', 'Theo tỉnh']:
    ws = wb[sname]
    print(f'\n=== {sname}: rows={ws.max_row} cols={ws.max_column} ===')
    for i in range(1, min(15, ws.max_row+1)):
        row_data = []
        for j in range(1, ws.max_column+1):
            v = ws.cell(i,j).value
            if v and str(v) not in ['#N/A', '#DIV/0!']:
                row_data.append(f'C{j}={str(v)[:50]}')
        if row_data: print(f'  R{i}: {row_data}')

# Read DOCX more thoroughly - especially about roles and KPIs
from docx import Document
docx_files = [f for f in os.listdir(folder) if f.endswith('.docx')]
docx_path = os.path.join(folder, docx_files[0])
doc = Document(docx_path)
print('\n\n=== DOCX - ALL content ===')
for i, p in enumerate(doc.paragraphs):
    if p.text.strip():
        print(f'{p.text}')
